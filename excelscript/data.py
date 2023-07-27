import logging
import pathlib
from copy import copy

import openpyxl
import yaml
from openpyxl.styles import Font, Color, Alignment
from openpyxl.workbook import Workbook
from openpyxl.worksheet.cell_range import CellRange

logger = logging.getLogger(__name__)


def read_config():
    with open(pathlib.Path(__file__).parent / 'config.yml', encoding='utf-8') as f:
        config = yaml.load(f, Loader=yaml.FullLoader)
        return config


class DataHolder:
    def __init__(self, file, wb, config):
        self.file = file
        self.wb = wb
        self.config = config
        self.sheet_detail = {}
        for ws in wb:
            self.sheet_detail[ws.title] = {}
            self.sheet_detail[ws.title]['title_row1'] = 1
            self.sheet_detail[ws.title]['output'] = False
            cell_rc = self.search_keyword(ws['A1:BZ10'], self.config['分组'])
            self.sheet_detail[ws.title]['title_column2'] = 1
            if cell_rc:
                self.sheet_detail[ws.title]['title_row2'] = cell_rc[1]
                cell = ws.cell(row=cell_rc[1], column=cell_rc[0])
                self.sheet_detail[ws.title]['key_cell'] = cell.coordinate
                logger.debug(f'sheet({ws.title})的分组单元格为{cell.coordinate}, {cell.value}')
            else:
                self.sheet_detail[ws.title]['title_row2'] = 1
                cell = ws.cell(row=1, column=1)
                self.sheet_detail[ws.title]['key_cell'] = cell.coordinate
                logger.debug(f'sheet({ws.title})的分组单元格未找到,使用默认值 {cell.coordinate}, {cell.value}')
            # 方便测试
            # self.sheet_detail[ws.title]['title_row2'] = 3

    @staticmethod
    def create(file):
        wb = openpyxl.load_workbook(file, rich_text=True, data_only=True)
        config = read_config()
        return DataHolder(file, wb, config)

    def search_keyword(self, area, keyword):
        for row in area:
            for cell in row:
                if type(cell.value) == str and keyword == cell.value.strip():
                    return cell.column, cell.row

    def gen(self, progress_callback):
        path = pathlib.Path(self.config['输出'])
        path.mkdir(parents=True, exist_ok=True)

        header_excel = pathlib.Path(self.config['输出']) / 'header.xlsx'
        self.gen_header(header_excel, progress_callback)
        return self.gen_excel(header_excel, progress_callback)

    def gen_header(self, header_excel, progress_callback):
        wb2 = Workbook()
        wb2.remove(wb2.active)

        progress_callback('正在生成表头..')

        for ws in self.wb:
            if not self.sheet_detail[ws.title]['output']:
                continue
            ws_config = self.sheet_detail[ws.title]

            ws2 = wb2.create_sheet(ws.title)

            title_area = f"A{ws_config['title_row1']}:BZ{ws_config['title_row2']}"
            bound = openpyxl.utils.cell.range_boundaries(title_area)

            # 合并单元格
            area = CellRange(title_area)
            for mcr in ws.merged_cells:
                if area.isdisjoint(mcr):
                    continue
                cr = CellRange(mcr.coord)
                ws2.merge_cells(cr.coord)

            # 复制单元格格式
            for row in ws[title_area]:
                for cell in row:
                    if cell.value and cell.column > ws_config['title_column2']:
                        ws_config['title_column2'] = cell.column
                    dst_cell = ws2.cell(row=cell.row, column=cell.column)
                    copy_cell(cell, dst_cell)

            ws2.sheet_format = copy(ws.sheet_format)
            ws2.page_margins = copy(ws.page_margins)

            # 设置列宽度
            for i in range(bound[0], bound[2] + 1):
                column_letter = openpyxl.utils.get_column_letter(i)
                if column_letter in ws.column_dimensions:
                    ws2.column_dimensions[column_letter].width = ws.column_dimensions[column_letter].width

            # 设置行宽度
            for i in range(bound[1], bound[3] + 1):
                if i in ws.row_dimensions:
                    ws2.row_dimensions[i].height = ws.row_dimensions[i].height

            # 拆分单元格
            merged_cells_columns = set()
            for mcr in ws.merged_cells:
                mcr_bound = openpyxl.utils.cell.range_boundaries(mcr.coord)
                if bound[3] < mcr_bound[1] < mcr_bound[3] and mcr_bound[0] == mcr_bound[2]:
                    if mcr_bound[0] not in merged_cells_columns:
                        merged_cells_columns.add(mcr_bound[0])
            ws_config['merged_cells_columns'] = merged_cells_columns

        wb2.save(header_excel)
        logger.info(f'生成表头成功:{self.sheet_detail}')

    def gen_excel(self, header_excel, progress_callback):
        progress_callback('解析excel..')
        save_workbooks = []
        fp_mapping = {}
        cfg = self.config['导出']
        for k in cfg:
            map_list = cfg[k]['映射']
            out_excel = k + '.xlsx'

            wb2 = openpyxl.load_workbook(header_excel, rich_text=True, data_only=True)
            fp = {'out': out_excel, 'wb': wb2, 'row': 0, 'dirty': False}
            save_workbooks.append(fp)

            for m in map_list:
                if m not in fp_mapping:
                    fp_mapping[m] = []
                fp_mapping[m].append(fp)

        totalSheetCount = 0
        finishedSheetCount = 0
        for ws in self.wb:
            if self.sheet_detail[ws.title]['output']:
                totalSheetCount += 1

        notClassified = set()

        for ws in self.wb:
            if not self.sheet_detail[ws.title]['output']:
                continue

            logger.info(f'拆分表格:{ws.title}')
            progress_callback(f'拆分表格{finishedSheetCount}/{totalSheetCount}')

            ws_cfg = self.sheet_detail[ws.title]

            row = ws_cfg['title_row2'] + 1
            col = ws[ws_cfg['key_cell']].column
            for swb in save_workbooks:
                swb['row'] = row
                swb['wb'].active = swb['wb'][ws.title]

            last_row = {}
            merged_cells_columns = ws_cfg['merged_cells_columns']

            while True:
                cell = ws.cell(row=row, column=col)
                if type(cell) != openpyxl.cell.cell.MergedCell:
                    if cell.value is None:
                        break
                    if type(cell.value) == str:
                        s = cell.value.strip()
                        if s in fp_mapping:
                            for swb in fp_mapping[s]:
                                copy_line(ws, swb, row, 1, ws_cfg['title_column2'], merged_cells_columns,
                                          last_row)
                        elif s not in self.config['过滤']:
                            notClassified.add(s)
                            logger.info(f'未归类的分组: {ws.title}, {s}')
                row = row + 1
            finishedSheetCount += 1

        wbTotal = len(save_workbooks)
        wbCount = 0
        # 保存
        for swb in save_workbooks:
            progress_callback(f'生成excel {wbCount}/{wbTotal}..')
            out_file = pathlib.Path(self.config['输出']) / swb['out']
            if swb['dirty']:
                swb['wb'].save(out_file)
                logger.info(f'保存excel:{out_file}')
            else:
                logger.info(f'表格为空，已过滤:{out_file}')
            wbCount += 1
        logger.info('导出excel完成!')
        return notClassified


def copy_cell(src_cell, dst_cell):
    if type(src_cell) != openpyxl.cell.cell.MergedCell:
        dst_cell.value = src_cell.value
        dst_cell.data_type = src_cell.data_type

    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = copy(src_cell.number_format)
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)


def fast_copy_font(src_cell, dst_cell):
    src_color = src_cell.font.color
    if src_color is None:
        return
    if src_color.type == 'rgb':
        dst_cell.font = Font(color=Color(rgb=src_color.rgb))


def copy_line(ws, swb, row, min_col, max_col, merged_cells_columns, last_row):
    ws2 = swb['wb'].active
    for col in range(min_col, max_col + 1):
        src_cell = ws.cell(row=row, column=col)
        dst_cell = ws2.cell(row=swb['row'], column=col)

        if type(src_cell) == openpyxl.cell.cell.MergedCell and col in merged_cells_columns:
            target_cell = last_row[col]
        else:
            target_cell = last_row[col] = src_cell
        dst_cell.value = target_cell.value
        dst_cell.number_format = target_cell.number_format
        fast_copy_font(target_cell, dst_cell)
        dst_cell.alignment = Alignment(horizontal=target_cell.alignment.horizontal,
                                       vertical=target_cell.alignment.vertical)
        # 很慢
        # copy_cell(target_cell, dst_cell)

    swb['row'] = swb['row'] + 1
    swb['dirty'] = True
